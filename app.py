import datetime
import os
import re
import sqlite3
import sys
from openpyxl import load_workbook
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, abort
from flask_login import current_user, LoginManager, login_user, logout_user, login_required, UserMixin
from urllib.parse import urlsplit
from werkzeug.utils import secure_filename


app = Flask(__name__)
app.secret_key = "your_secret_key"  # Change this to a secure secret key

login_manager = LoginManager()
login_manager.login_view  = 'login'
login_manager.init_app(app)
app.config['USERNAME'] = ""
app.config['PASSWORD'] = ""


class FakeUser(UserMixin):
    id = 1
    username = "Admin User"
    password = ""

@login_manager.user_loader
def load_user(id):
    fakeuser = FakeUser()
    return fakeuser



UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['DATABASE'] = 'db/data.db'
app.config['AID_STATION_MAP'] = {
    "AS1": "Aid Station 1", 
    "AS2": "Aid Station 2", 
    "AS3": "Aid Station 3", 
    "AS46": "Aid Station 4/6", 
    "AS5": "Aid Station 5", 
    "AS7": "Aid Station 7", 
    "AS8": "Aid Station 8", 
    "AS9": "Aid Station 9", 
    "AS10": "Aid Station 10", 
    "mA": "Med Alpha", 
    "mB": "Med Bravo", 
    "mC": "Med Charlie", 
    "mD": "Med Delta", 
    "mE": "Med Echo"
}
app.config['BOOLEAN_MAP'] = {
    0: " ",
    1: "Yes"
}

# This should be a recursive walk for the database path... TODO
if not os.path.exists('db'):
    os.makedirs('db')


# Function to export data as a zipped dict
def zip_data(cursor, table_name, id=None):
    if id is None:
        cursor.execute(f"SELECT * FROM {table_name}")
    else:
        cursor.execute(f"SELECT * FROM {table_name} WHERE ID={id}")
    rows = cursor.fetchall()


    # Get the column names
    cursor.execute(f"PRAGMA table_info({table_name})")
    columns = [column[1] for column in cursor.fetchall()]

    # Convert the data to a list of dictionaries
    data_list = []
    for row in rows:
        data_dict = dict(zip(columns, row))
        data_list.append(data_dict)

    return {'data': data_list}



# Function to fetch a sqlite table as a JSON string
def load_data(table_name, id=None):
    # Connect to the SQLite database
    conn = db_connect()
    cursor = conn.cursor()
    data = zip_data(cursor, table_name, id)
    conn.close()
    return data

# Function to check if the file extension is allowed
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Function to connect to SQLLite Database
def db_connect():
    return sqlite3.connect(app.config['DATABASE'])

# Function to create an SQLite database and table to store data
def create_database():
    conn =  db_connect()
    cursor = conn.cursor()
    cursor.execute('''CREATE TABLE IF NOT EXISTS encounters (
                      id INTEGER PRIMARY KEY AUTOINCREMENT,
                      aid_station TEXT,
                      bib TEXT,
                      report_time TEXT,
                      discharged INTEGER DEFAULT 0 NOT NULL,
                      transported INTEGER DEFAULT 0 NOT NULL,
                      hospital TEXT
                   )''')

    cursor.execute('''CREATE TABLE IF NOT EXISTS notes (
                      id INTEGER PRIMARY KEY AUTOINCREMENT,
                      aid_station TEXT,
                      note TEXT,
                      report_time TEXT
                   )''')
    print("Database created!", file=sys.stderr)
    conn.commit()
    conn.close()



# *****************************************************************************
@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('upload_file'))

    # If a post request was made, find the user by 
    # filtering for the username
    if request.method == "POST":
        if app.config['USERNAME'] == request.form.get("username"):
            if app.config['PASSWORD'] == request.form.get("password"):
                user = FakeUser()
                user.username = app.config['USERNAME'] 
                user.password = app.config['PASSWORD']
                # user.user_id = 1;
                login_user(user, remember='y')
                next_page = request.args.get('next')
                if not next_page or urlsplit(next_page).netloc != '':
                    next_page = url_for('upload_file')
                return redirect(next_page)
        # Redirect the user back to the home
        # (we'll create the home route in a moment)
    return render_template("login.html")

@app.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('login'))


# Route for uploading the XLSX file
@app.route('/', methods=['GET', 'POST'])
@login_required
def upload_file():
    if request.method == 'POST':
        
        # Validate the post request
        errors = []
        if 'file' not in request.files:
            errors.append('Please select a file and try again.')

        if 'aidstation' not in request.form:
            errors.append('Please select an Aid Station and try again.')

        if errors:
            for error in errors:
                flash(error, 'error')
            return redirect(request.url)

        aidstation = request.form['aidstation']
        file = request.files['file']

        # Validate form contents
        if aidstation == '':
            errors.append('Please select a file and try again.')

        if file.filename == '':
            errors.append('Please select an Aid Station and try again.')

        if errors:
            for error in errors:
                flash(error, 'error')
            return redirect(request.url)


        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)

            # Get the current date and time
            current_time = datetime.datetime.now()

            # Format the date and time as a string
            timestamp = current_time.strftime("%Y-%m-%d_%H-%M-%S")

            # Add timestamp and aidstaiton to filename
            filename = f"{aidstation}_{timestamp}_{filename}"

            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))

            # Parse the uploaded XLSX file and store data in the database
            
            parse_and_store_data(os.path.join(app.config['UPLOAD_FOLDER'], filename), aidstation)

            flash('File successfully uploaded and data stored in the database', 'success')
            return redirect(request.url)
        else:
            flash('File type is not allowed, please chose and Excel (xlsx) document.', 'error')

    conn = db_connect();
    cursor = conn.cursor()

    # Retrieve encounters
    cursor.execute("SELECT * FROM encounters ORDER BY aid_station, report_time DESC")
    encounters = cursor.fetchall()

    # # Retrieve data from the notes table
    # cursor.execute("SELECT * FROM notes ORDER BY aid_station, report_time DESC")
    # notes = cursor.fetchall()

    # Retrieve a list of all the Aid Stations 
    cursor.execute("SELECT DISTINCT aid_station FROM encounters")
    aid_stations = cursor.fetchall();

    active_encounters_by_station = {}
    synopsis = {}
    # all_encounters_by_station = {}
    synopsis['total'] = {}
    synopsis['stations'] = {}

    cursor.execute("SELECT COUNT(*) FROM encounters")
    synopsis['total']['encounters'] = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM encounters WHERE discharged IS NOT 1")
    synopsis['total']['active'] = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM encounters WHERE discharged=1")
    synopsis['total']['discharged'] = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM encounters WHERE transported=1")
    synopsis['total']['transported'] = cursor.fetchone()[0]


    for aid_station in aid_stations:
        (aid_station_str,) = aid_station
        cursor.execute("SELECT * FROM encounters WHERE discharged IS NOT 1 AND aid_station=?", (aid_station))
        active_encounters_by_station[aid_station_str] = cursor.fetchall()


    for aid_station in app.config['AID_STATION_MAP'].keys():
        synopsis['stations'][aid_station] = {}
        cursor.execute("SELECT COUNT(*) FROM encounters WHERE aid_station=?", (aid_station,))
        synopsis['stations'][aid_station]['encounters'] = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM encounters WHERE discharged IS NOT 1 AND aid_station=?", (aid_station,))
        synopsis['stations'][aid_station]['active'] = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM encounters WHERE discharged=1 AND aid_station=?", (aid_station,))
        synopsis['stations'][aid_station]['discharged'] = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM encounters WHERE transported=1 AND aid_station=?", (aid_station,))
        synopsis['stations'][aid_station]['transported'] = cursor.fetchone()[0]

        # cursor.execute("SELECT * FROM encounters WHERE aid_station=?", (aid_station))
        # all_encounters_by_station[aid_station_str] = cursor.fetchall()

    # Retrieve all active encounters grouped by aid station
    # cursor.execute("SELECT * FROM encounters WHERE discharged=0 GROUP BY aid_station")
    # active_grouped = cursor.fetchall()

    # Close the database connection
    conn.close()

    return render_template('upload.html',
                            aid_station_map=app.config['AID_STATION_MAP'],
                            encounters=encounters,
                            aid_stations=aid_stations,
                            synopsis=synopsis,
                            active_encounters=active_encounters_by_station)


@app.route('/files/')
@login_required
def list_uploads():

    # Filter only xlxs
    file_list = [f for f in os.listdir(app.config['UPLOAD_FOLDER']) if f.lower().endswith('.xlsx')]
    sorted_file_list = sorted(file_list, reverse=True)

    file_list_by_aid = []

    for asdx in app.config['AID_STATION_MAP'].keys():
        print(f"In aid station {asdx}", file=sys.stderr)
        file_list_by_aid.append({
            "name": app.config['AID_STATION_MAP'][asdx],
            "files": [f for f in sorted_file_list if f.upper().startswith(asdx)]
        })


    return render_template('list_uploads.html', files=file_list_by_aid)


@app.route('/download/<file_name>')
@login_required
def download_file(file_name):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], file_name)
    if os.path.isfile(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        abort(404)

@app.route('/encounters', methods=['GET'])
@login_required
def get_encounters():
    return render_template('encounters.html')

@app.route('/notes', methods=['GET'])
@login_required
def get_notes():
    return render_template('notes.html')


###############################################################################
@app.route('/api/encounters', methods=['GET', 'POST'])
@login_required
def api_encounters():
    if request.method == 'POST':
        
        # Validate the post request
        if 'action' not in request.form:
            return jsonify({ 'error': 'Ahhh I dont know what to do, please provide an action'})

        action = request.form['action']

        pattern = r'\[(\d+)\]\[([a-zA-Z_]+)\]'
        data = {}
        id = 0

        for key in request.form.keys():
            print(f"Key: {key}", file=sys.stderr)
            matches = re.search(pattern, key)
            if matches:
                id = int(matches.group(1))
                field_key = matches.group(2)
                data[field_key] = request.form[key]

        
        # Handle Editing an existing record
        if action.lower() == 'edit':
            conn = db_connect()
            cursor = conn.cursor()
            cursor.execute('''UPDATE encounters
                            SET aid_station=?, bib=?, report_time=?, discharged=?, transported=?, hospital=?
                            WHERE id=?
                           ''', (data['aid_station'], data['bib'], data['report_time'], data['discharged'], data['transported'], data['hospital'], id))
            new_data = zip_data(cursor, 'encounters', id)
            conn.commit()
            conn.close()
            return jsonify(new_data)

        # Handle Creating a new record
        if action.lower() == 'create':
            conn = db_connect()
            cursor = conn.cursor()
            cursor.execute('''INSERT INTO encounters
                            (aid_station, bib, report_time, discharged, transported, hospital)
                            VALUES (?, ?, ?, ?, ?, ?)
                           ''', (data['aid_station'], data['bib'], data['report_time'], data['discharged'], data['transported'], data['hospital']))
            
            new_data = zip_data(cursor, 'encounters', cursor.lastrowid)
            conn.commit()
            conn.close()
            return jsonify(new_data)


        # Handle Remove
        if action.lower() == 'remove':
            conn = db_connect()
            cursor = conn.cursor()
            cursor.execute(f"DELETE FROM encounters WHERE id={id}")
            conn.commit()
            
            new_data = zip_data(cursor, 'encounters')
            conn.close()
            return jsonify(new_data)

    return jsonify(load_data('encounters'))

###############################################################################
@app.route('/api/notes', methods=['GET', 'POST'])
@login_required
def api_notes():
    if request.method == 'POST':
        
        # Validate the post request
        if 'action' not in request.form:
            return jsonify({ 'error': 'Ahhh I dont know what to do, please provide an action'})

        action = request.form['action']

        pattern = r'\[(\d+)\]\[([a-zA-Z_]+)\]'
        data = {}
        id = 0

        for key in request.form.keys():
            print(f"Key: {key}", file=sys.stderr)
            matches = re.search(pattern, key)
            if matches:
                id = int(matches.group(1))
                field_key = matches.group(2)
                data[field_key] = request.form[key]

        
        # Handle Editing an existing record
        if action.lower() == 'edit':
            conn = db_connect()
            cursor = conn.cursor()
            cursor.execute('''UPDATE notes
                            SET aid_station=?, note=?, report_time=?
                            WHERE id=?
                           ''', (data['aid_station'], data['note'], data['report_time'], id))
            new_data = zip_data(cursor, 'notes', id)
            conn.commit()
            conn.close()
            return jsonify(new_data)

        # Handle Creating a new record
        if action.lower() == 'create':
            conn = db_connect()
            cursor = conn.cursor()
            cursor.execute('''INSERT INTO notes
                            (aid_station, note, report_time)
                            VALUES (?, ?, ?)
                           ''', (data['aid_station'], data['note'], data['report_time']))
            
            new_data = zip_data(cursor, 'notes', cursor.lastrowid)
            conn.commit()
            conn.close()
            return jsonify(new_data)


        # Handle Remove
        if action.lower() == 'remove':
            conn = db_connect()
            cursor = conn.cursor()
            cursor.execute(f"DELETE FROM notes WHERE id={id}")
            conn.commit()
            
            new_data = zip_data(cursor, 'notes')
            conn.close()
            return jsonify(new_data)

    return jsonify(load_data('notes'))

# Function to check if this is a new encounter for the aid staiton and bib
def is_new_encounter(cursor, aidstation, bib):
    cursor.execute('''SELECT bib FROM encounters 
                      WHERE bib=? AND aid_station=?
                      ''', (bib, aidstation))

    gresult = cursor.fetchone()
    
    if gresult:
        return False
    else:
        return True


# Function to add a row to the current encounters table
def add_encounter(cursor, aidstation, bib, report_time, discharged=0, transported=0, hospital=None):
    # Check if we already have it
    if is_new_encounter(cursor, aidstation, bib):
        cursor.execute('''INSERT INTO encounters 
                                (aid_station, bib, report_time, discharged, transported, hospital) 
                                VALUES (?, ?, ?, ?, ?, ?)
                       ''', (aidstation, bib, report_time, discharged, transported, hospital))
    else:
        cursor.execute('''UPDATE encounters
                          SET discharged=?, transported=?, hospital=?
                          WHERE bib=? AND aid_station=?
                       ''', (discharged, transported, hospital, bib, aidstation))


# Function to handle notes
def add_note(cursor, aidstation, report_time, note):
    cursor.execute('''INSERT INTO notes
                        (aid_station, note, report_time) 
                        VALUES (?, ?, ?)
                   ''', (aidstation, note, report_time))


# Function to parse the XLSX file and store data in the database
def parse_and_store_data(file_path, aidstation):
    wb = load_workbook(file_path)
    sheet = wb["Census Roster Sheet"] # wb.active
    conn = db_connect();
    cursor = conn.cursor()

    # Get the Report time Value HHMM
    report_time = sheet["C4"].value

    
    # Encounter Ranges for Current Encounters
    encounter_ranges = [{'min': 7, 'max': 16}, {'min': 48, 'max': 57}]

    
    # Iterate over ranges of current encounters
    for ier in encounter_ranges:

        # Iterate over Bib numbers currently in Aid Station
        for row in sheet.iter_rows(values_only=True, min_row=ier['min'], max_col=2, max_row=ier['max']):

            # Iterate over each cell in the row and call it a bib
            for bib in row:

                # Skip if there is no value in cell
                if (bib != None):
                    add_encounter(cursor, aidstation, bib, report_time)
                    

    # Encounter Ranges for Previous Encounters
    encounter_ranges = [{'min': 19, 'max': 28}, {'min': 60, 'max': 69}]

    # Iterate over ranges for previous encounters
    for ier in encounter_ranges:

        # Iterate over Bib numbers discharged since last report
        for row in sheet.iter_rows(values_only=True, min_row=ier['min'], max_col=2, max_row=ier['max']):

            # Iterate over each cell in the row and call is a bib
            for bib in row:

                # Skip if there is no value in cell
                if (bib != None):
                    add_encounter(cursor, aidstation, bib, report_time, 1)

    # Encounter Ranges for Hospital Transport
    encounter_ranges = [{'min': 31, 'max': 35}, {'min': 72, 'max': 76}]

    # Iterate over ranges for transport
    for ier in encounter_ranges:

        # Iterate over Bib numbers discharged since last report
        for row in sheet.iter_rows(values_only=True, min_row=ier['min'], max_col=2, max_row=ier['max']):

            bib, hospital = row

            # Skip if there is no value in cell
            if (bib != None):
                add_encounter(cursor, aidstation, bib, report_time, 1, 1, hospital)

    # Encounter Ranges for notes
    encounter_ranges = [{'min': 38, 'max': 40}, {'min': 79, 'max': 81}]

    # Iterate over ranges for previous encounters
    for ier in encounter_ranges:

        # Iterate over Bib numbers discharged since last report
        for row in sheet.iter_rows(values_only=True, min_row=ier['min'], max_col=1, max_row=ier['max']):

            for note in row:

                # Skip if there is no value in cell
                if (note != None):
                    add_note(cursor, aidstation, report_time, note)



    conn.commit()
    conn.close()

if __name__ == '__main__':
    create_database()
    app.run(debug=True)
    
